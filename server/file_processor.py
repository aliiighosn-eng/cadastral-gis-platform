"""
File processing utilities for reading and exporting GIS data.
Handles .dxf, .geojson, and .tab file formats.
"""

import json
import os
from typing import Dict, List, Optional

import fiona
import geopandas as gpd
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from server.gis_utils import CoordinateTransformer, GeometryParser


class FileProcessor:
    """Processes GIS files in various formats."""

    SUPPORTED_FORMATS = {".geojson", ".dxf", ".tab", ".shp"}

    @staticmethod
    def read_geojson(file_path: str) -> Optional[gpd.GeoDataFrame]:
        """Read GeoJSON file."""
        try:
            return gpd.read_file(file_path)
        except (FileNotFoundError, ValueError, KeyError) as e:
            raise ValueError(f"Failed to read GeoJSON: {str(e)}")

    @staticmethod
    def read_dxf(file_path: str) -> Optional[gpd.GeoDataFrame]:
        """Read DXF file using fiona."""
        try:
            # DXF files may have multiple layers
            layers = fiona.listlayers(file_path)
            gdf_list = []

            for layer in layers:
                try:
                    gdf = gpd.read_file(file_path, layer=layer)
                    gdf_list.append(gdf)
                except (FileNotFoundError, ValueError, KeyError):
                    continue

            if gdf_list:
                return gpd.GeoDataFrame(pd.concat(gdf_list, ignore_index=True))
            return None
        except (FileNotFoundError, ValueError, KeyError) as e:
            raise ValueError(f"Failed to read DXF: {str(e)}")

    @staticmethod
    def read_tab(file_path: str) -> Optional[gpd.GeoDataFrame]:
        """Read TAB (MapInfo) file."""
        try:
            return gpd.read_file(file_path)
        except (FileNotFoundError, ValueError, KeyError) as e:
            raise ValueError(f"Failed to read TAB: {str(e)}")

    @staticmethod
    def read_file(file_path: str) -> Optional[gpd.GeoDataFrame]:
        """Read GIS file in any supported format."""
        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".geojson":
            return FileProcessor.read_geojson(file_path)
        elif ext == ".dxf":
            return FileProcessor.read_dxf(file_path)
        elif ext == ".tab":
            return FileProcessor.read_tab(file_path)
        elif ext in [".shp", ".gpkg"]:
            return gpd.read_file(file_path)
        else:
            raise ValueError(f"Unsupported file format: {ext}")

    @staticmethod
    def extract_features(gdf: gpd.GeoDataFrame) -> List[Dict]:
        """Extract features from GeoDataFrame."""
        features = []

        for idx, row in gdf.iterrows():
            feature = {
                "id": idx,
                "geometry": row.geometry.__geo_interface__,
                "properties": {
                    k: v for k, v in row.items() if k != "geometry"
                },
            }
            features.append(feature)

        return features

    @staticmethod
    def export_to_excel(
        features: List[Dict],
        output_path: str,
        target_system: str = "EPSG:4328",
        source_system: str = "EPSG:4326",
    ) -> str:
        """
        Export features to Excel with transformed coordinates.

        Args:
            features: List of feature dictionaries
            output_path: Path to save Excel file
            target_system: Target coordinate system
            source_system: Source coordinate system

        Returns:
            Path to created Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Coordinates"

        # Create header
        headers = ["ID", "kadNum", "GeometryType", "x", "y", "System"]
        ws.append(headers)

        # Style header
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Process features
        row_num = 2
        for feature in features:
            geometry = feature.get("geometry", {})
            properties = feature.get("properties", {})
            feature_id = feature.get("id", "")
            kad_num = properties.get("kadNumber", properties.get("kadNum", ""))
            geom_type = GeometryParser.get_geometry_type(geometry)

            # Extract coordinates
            coords = GeometryParser.extract_coordinates(geometry)

            for x, y in coords:
                # Transform coordinates if needed
                try:
                    transformed_x, transformed_y = (
                        CoordinateTransformer.transform_coordinates(
                            x, y, source_system, target_system
                        )
                    )
                except (ValueError, TypeError):
                    transformed_x, transformed_y = x, y

                ws.append(
                    [
                        feature_id,
                        kad_num,
                        geom_type,
                        round(transformed_x, 6),
                        round(transformed_y, 6),
                        target_system,
                    ]
                )
                row_num += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

        # Save file
        wb.save(output_path)
        return output_path


class GeoJSONMerger:
    """Merges multiple GeoJSON files into a single file."""

    @staticmethod
    def merge_layers(
        file_paths: List[str],
        output_path: str,
        target_system: str = "EPSG:4328",
    ) -> str:
        """
        Merge multiple GeoJSON layers into a single file.

        Args:
            file_paths: List of GeoJSON file paths
            output_path: Path to save merged file
            target_system: Target coordinate system

        Returns:
            Path to merged GeoJSON file
        """
        merged_features = []
        required_fields = ["ID", "kadNum", "name", "landUse"]

        for file_path in file_paths:
            try:
                gdf = gpd.read_file(file_path)

                for idx, row in gdf.iterrows():
                    feature = {
                        "type": "Feature",
                        "geometry": row.geometry.__geo_interface__,
                        "properties": {},
                    }

                    # Add required fields, using None for missing values
                    for field in required_fields:
                        feature["properties"][field] = row.get(field, None)

                    # Add all other properties
                    for key, value in row.items():
                        if key not in required_fields and key != "geometry":
                            feature["properties"][key] = value

                    merged_features.append(feature)
            except (FileNotFoundError, ValueError, KeyError) as e:
                print(f"Warning: Could not process {file_path}: {str(e)}")
                continue

        # Create FeatureCollection
        feature_collection = {
            "type": "FeatureCollection",
            "features": merged_features,
        }

        # Save to file
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(feature_collection, f, ensure_ascii=False, indent=2)

        return output_path

    @staticmethod
    def validate_geojson(file_path: str) -> bool:
        """Validate GeoJSON file structure."""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            if data.get("type") not in ["FeatureCollection", "Feature"]:
                return False

            if data.get("type") == "FeatureCollection":
                if not isinstance(data.get("features", []), list):
                    return False

            return True
        except (FileNotFoundError, json.JSONDecodeError, ValueError):
            return False
